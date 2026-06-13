# -*- coding: utf-8 -*-
"""
سكربت إثراء الكتالوج — يُنفّذ مرة واحدة
يقرأ ملف Excel الشامل ويستخلص التصنيف والمكونات والعائلة العطرية
ثم يحدّث catalog.json بالبيانات الجديدة

Usage:
    python enrich_catalog.py
"""

import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

import json
import re
import os
from pathlib import Path

# ═══════════════════════════════════════════════════════════
#  إعدادات
# ═══════════════════════════════════════════════════════════

BASE_DIR = Path(__file__).parent
EXCEL_FILE = Path(r'C:\Users\Hp\Downloads\mahwous_08-06-2026-18-27_kPTl4zMgsT9vxbcCmNMzUkeOh4kEMRwRr5n3fE8k_products.xlsx')
CATALOG_FILE = BASE_DIR / 'catalog.json'
BACKUP_FILE = BASE_DIR / 'catalog_backup.json'


# ═══════════════════════════════════════════════════════════
#  عوائل الروائح — الكلمات المفتاحية لكل عائلة
# ═══════════════════════════════════════════════════════════

SCENT_FAMILY_KEYWORDS = {
    'oud': [
        'عود', 'العود', 'أود', 'oud', 'aoud', 'دهن العود',
        'عود كمبودي', 'عود هندي', 'أجار', 'agarwood',
    ],
    'oriental': [
        'شرقي', 'شرقية', 'oriental', 'عنبر', 'amber', 'بخور',
        'لبان', 'عبير', 'تابلي', 'spicy', 'زعفران', 'saffron',
        'قرفة', 'cinnamon', 'هيل', 'cardamom', 'جوزة الطيب', 'nutmeg',
        'قرنفل', 'clove', 'فلفل', 'pepper',
    ],
    'woody': [
        'خشبي', 'خشبية', 'woody', 'أخشاب', 'خشب الصندل', 'sandalwood',
        'خشب الأرز', 'cedarwood', 'cedar', 'باتشولي', 'patchouli',
        'فيتيفر', 'vetiver', 'غايك', 'guaiac',
    ],
    'floral': [
        'زهري', 'زهرية', 'floral', 'ورد', 'rose', 'ياسمين', 'jasmine',
        'فاوانيا', 'peony', 'زنبق', 'lily', 'توبيروز', 'tuberose',
        'إيلنغ', 'ylang', 'بنفسج', 'violet', 'زهر البرتقال', 'neroli',
        'لافندر', 'lavender', 'ماغنوليا', 'magnolia', 'أوركيد', 'orchid',
    ],
    'fresh': [
        'فريش', 'منعش', 'منعشة', 'fresh', 'أكواتيك', 'aquatic',
        'بحري', 'marine', 'أوزون', 'ozone', 'نظيف', 'clean',
        'مائي', 'شلال', 'نسيم',
    ],
    'citrus': [
        'حمضي', 'حمضية', 'citrus', 'ليمون', 'lemon', 'برغموت', 'bergamot',
        'برتقال', 'orange', 'جريب فروت', 'grapefruit', 'يوسفي', 'mandarin',
        'ليم', 'lime', 'كباد',
    ],
    'sweet': [
        'سويت', 'حلو', 'sweet', 'فانيلا', 'vanilla', 'كراميل', 'caramel',
        'توفي', 'toffee', 'شوكولاتة', 'chocolate', 'عسل', 'honey',
        'حلوى', 'سكر', 'sugar', 'بودرة', 'powder', 'بودري',
        'كيك', 'كيكة', 'لاتيه', 'latte', 'بسكويت', 'marshmallow',
    ],
    'musk': [
        'مسك', 'مسكي', 'musk', 'المسك', 'مسك أبيض', 'white musk',
        'مسك طهارة', 'مسك الختام',
    ],
    'leather': [
        'جلد', 'جلود', 'leather', 'سويد', 'suede', 'جلد محروق',
    ],
    'gourmand': [
        'قهوة', 'coffee', 'كاكاو', 'cocoa', 'حليب', 'milk',
        'لوز', 'almond', 'فستق', 'pistachio', 'فراولة', 'strawberry',
        'توت', 'berry', 'خوخ', 'peach', 'تفاح', 'apple',
        'مانجو', 'mango', 'كوكو', 'coconut',
    ],
}

# أولوية العوائل (الأهم أولاً) — لو المنتج ينتمي لأكثر من عائلة
FAMILY_PRIORITY = ['oud', 'oriental', 'woody', 'leather', 'floral', 'sweet',
                   'gourmand', 'citrus', 'fresh', 'musk']


# ═══════════════════════════════════════════════════════════
#  تصنيف الجنس من اسم التصنيف
# ═══════════════════════════════════════════════════════════

def detect_gender_from_category(category_str):
    """استنتاج الجنس من تصنيف المنتج"""
    if not category_str:
        return None
    cat_lower = category_str.lower()
    has_male = any(kw in cat_lower for kw in ['رجالي', 'رجاليه', 'للرجال', 'men', 'male', 'رجالية'])
    has_female = any(kw in cat_lower for kw in ['نسائي', 'نسائيه', 'للنساء', 'women', 'female', 'نسائية'])
    if has_male and has_female:
        return 'مشترك'
    if has_male:
        return 'رجالي'
    if has_female:
        return 'نسائي'
    # مشترك/للجنسين
    if any(kw in cat_lower for kw in ['للجنسين', 'unisex', 'مشترك']):
        return 'مشترك'
    return None


# ═══════════════════════════════════════════════════════════
#  استخلاص المكونات من الوصف HTML
# ═══════════════════════════════════════════════════════════

def strip_html(html_text):
    """إزالة وسوم HTML وإرجاع نص نظيف"""
    if not html_text:
        return ''
    text = re.sub(r'<br\s*/?>', '\n', html_text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '\n', text)
    text = re.sub(r'&nbsp;', ' ', text)
    text = re.sub(r'&amp;', '&', text)
    text = re.sub(r'&[a-zA-Z]+;', '', text)
    text = re.sub(r'\n+', '\n', text)
    return text.strip()


def extract_ingredients_from_desc(desc_html):
    """استخلاص المكونات العطرية من الوصف.

    يبحث عن أقسام الافتتاحية/القلب/القاعدة ويستخرج النوتات المذكورة.
    يرجع dict: {'top': '...', 'heart': '...', 'base': '...', 'combined': '...'}
    """
    clean = strip_html(desc_html)
    if not clean:
        return {'top': '', 'heart': '', 'base': '', 'combined': ''}

    lines = [l.strip() for l in clean.split('\n') if l.strip()]

    # أنماط بحث عن العناوين
    top_patterns = [
        r'الافتتاحية', r'الانطباع الأول', r'المقدمة', r'النوتات العليا',
        r'نوتات القمة', r'top\s*notes?', r'مقدمة العطر',
    ]
    heart_patterns = [
        r'القلب', r'قلب العطر', r'روح العطر', r'النوتات الوسطى',
        r'heart\s*notes?', r'middle\s*notes?',
    ]
    base_patterns = [
        r'القاعدة', r'قاعدة العطر', r'البصمة', r'النوتات الأساسية',
        r'الختام', r'base\s*notes?', r'الأساس',
    ]

    def _find_section(patterns, lines):
        """يبحث عن عنوان القسم ويرجع النص التالي له"""
        for i, line in enumerate(lines):
            for pat in patterns:
                if re.search(pat, line, re.IGNORECASE):
                    # النص التالي هو المحتوى
                    if i + 1 < len(lines):
                        content = lines[i + 1]
                        # تجاهل لو السطر التالي عنوان آخر
                        if len(content) > 15 and not any(
                            re.search(p, content, re.IGNORECASE)
                            for pats in [top_patterns, heart_patterns, base_patterns]
                            for p in pats
                        ):
                            return content
        return ''

    top = _find_section(top_patterns, lines)
    heart = _find_section(heart_patterns, lines)
    base = _find_section(base_patterns, lines)

    # إذا ما لقينا أقسام واضحة، نحاول استخلاص من الوصف العام
    combined = ''
    if top or heart or base:
        parts = [p for p in [top, heart, base] if p]
        combined = ' | '.join(parts)
    else:
        # بحث عن كلمات المكونات في النص الكامل
        ingredient_words = []
        all_keywords = []
        for family_keywords in SCENT_FAMILY_KEYWORDS.values():
            all_keywords.extend(family_keywords)

        for line in lines[:25]:  # أول 25 سطر فقط
            for kw in all_keywords:
                if kw in line.lower() and len(kw) > 2:
                    ingredient_words.append(kw)

        if ingredient_words:
            combined = '، '.join(list(dict.fromkeys(ingredient_words))[:10])

    return {
        'top': top[:200] if top else '',
        'heart': heart[:200] if heart else '',
        'base': base[:200] if base else '',
        'combined': combined[:400] if combined else '',
    }


# ═══════════════════════════════════════════════════════════
#  تحديد العائلة العطرية
# ═══════════════════════════════════════════════════════════

def detect_scent_family(name, category, ingredients_text, desc_text=''):
    """تحديد العائلة العطرية بناءً على الاسم والتصنيف والمكونات.

    يرجع العائلة الأكثر تطابقاً أو 'oriental' كافتراضي.
    """
    # نص البحث = اسم + تصنيف + مكونات + أول 300 حرف من الوصف
    search_text = f"{name} {category} {ingredients_text} {desc_text[:300]}".lower()

    scores = {}
    for family, keywords in SCENT_FAMILY_KEYWORDS.items():
        score = 0
        for kw in keywords:
            if kw.lower() in search_text:
                # كلمات أطول = ثقة أعلى
                weight = 2 if len(kw) > 4 else 1
                score += weight
        scores[family] = score

    # ترتيب حسب الأولوية عند التساوي
    best_family = 'oriental'  # افتراضي
    best_score = 0
    for family in FAMILY_PRIORITY:
        if scores.get(family, 0) > best_score:
            best_score = scores[family]
            best_family = family

    return best_family


# ═══════════════════════════════════════════════════════════
#  تنظيف المكونات لنص مختصر
# ═══════════════════════════════════════════════════════════

# كلمات المكونات العطرية المعروفة — مع الأشكال المتعددة
# كل tuple = (الاسم المعروض, [أشكال البحث])
KNOWN_INGREDIENTS = [
    # خشب
    ('عود', ['عود', 'العود', 'أود', 'oud', 'agarwood']),
    ('خشب الصندل', ['خشب الصندل', 'الصندل', 'صندل', 'sandalwood']),
    ('خشب الأرز', ['خشب الأرز', 'الأرز', 'أرز', 'cedarwood', 'cedar']),
    ('باتشولي', ['باتشولي', 'الباتشولي', 'patchouli']),
    ('فيتيفر', ['فيتيفر', 'الفيتيفر', 'vetiver']),
    # زهور
    ('ورد', ['ورد', 'الورد', 'روز', 'rose', 'ورد دمشقي', 'الورد الدمشقي', 'ورد تركي']),
    ('ياسمين', ['ياسمين', 'الياسمين', 'jasmine']),
    ('فاوانيا', ['فاوانيا', 'الفاوانيا', 'peony']),
    ('زنبق', ['زنبق', 'الزنبق', 'lily']),
    ('توبيروز', ['توبيروز', 'التيوبروز', 'تيوبروز', 'tuberose']),
    ('إيلنغ إيلنغ', ['إيلنغ', 'الإيلنغ', 'ylang']),
    ('بنفسج', ['بنفسج', 'البنفسج', 'violet']),
    ('زهر البرتقال', ['زهر البرتقال', 'neroli', 'نيرولي']),
    ('لافندر', ['لافندر', 'اللافندر', 'خزامى', 'lavender']),
    ('ماغنوليا', ['ماغنوليا', 'magnolia']),
    ('أوركيد', ['أوركيد', 'الأوركيد', 'orchid']),
    # حمضيات
    ('برغموت', ['برغموت', 'البرغموت', 'bergamot']),
    ('ليمون', ['ليمون', 'الليمون', 'lemon']),
    ('برتقال', ['برتقال', 'البرتقال', 'orange']),
    ('جريب فروت', ['جريب فروت', 'grapefruit']),
    ('يوسفي', ['يوسفي', 'اليوسفي', 'mandarin']),
    ('ليم', ['ليم', 'الليم', 'lime']),
    # توابل
    ('زعفران', ['زعفران', 'الزعفران', 'saffron']),
    ('قرفة', ['قرفة', 'القرفة', 'cinnamon']),
    ('هيل', ['هيل', 'الهيل', 'هال', 'cardamom']),
    ('جوزة الطيب', ['جوزة الطيب', 'nutmeg']),
    ('قرنفل', ['قرنفل', 'القرنفل', 'clove']),
    ('فلفل', ['فلفل أسود', 'فلفل وردي', 'الفلفل', 'pepper']),
    ('زنجبيل', ['زنجبيل', 'الزنجبيل', 'ginger']),
    ('توابل', ['توابل', 'التوابل', 'تابلي', 'spicy', 'spice']),
    # حلو
    ('فانيلا', ['فانيلا', 'الفانيلا', 'فانيليا', 'الفانيليا', 'vanilla']),
    ('كراميل', ['كراميل', 'الكراميل', 'caramel']),
    ('عسل', ['عسل', 'العسل', 'honey']),
    ('شوكولاتة', ['شوكولاتة', 'الشوكولاتة', 'chocolate', 'كاكاو']),
    ('توفي', ['توفي', 'toffee']),
    ('بودرة', ['بودرة', 'البودرة', 'بودري', 'powder']),
    # فواكه
    ('تفاح', ['تفاح', 'التفاح', 'apple']),
    ('خوخ', ['خوخ', 'الخوخ', 'peach']),
    ('توت', ['توت', 'التوت', 'berry', 'berries', 'توت أحمر']),
    ('فراولة', ['فراولة', 'الفراولة', 'strawberry']),
    ('مانجو', ['مانجو', 'المانجو', 'mango']),
    ('أناناس', ['أناناس', 'الأناناس', 'pineapple']),
    ('كمثرى', ['كمثرى', 'الكمثرى', 'إجاص', 'pear']),
    ('رمان', ['رمان', 'الرمان', 'pomegranate']),
    ('فواكه', ['فواكه حمراء', 'فواكه استوائية', 'فاكهي', 'فواكه', 'fruity']),
    # مسك وعنبر
    ('مسك', ['مسك', 'المسك', 'مسك أبيض', 'musk']),
    ('عنبر', ['عنبر', 'العنبر', 'amber', 'أمبر']),
    # جلد
    ('جلود', ['جلد', 'جلود', 'الجلد', 'الجلود', 'leather', 'سويد', 'suede']),
    # أخرى
    ('لبان', ['لبان', 'اللبان', 'frankincense']),
    ('بخور', ['بخور', 'البخور', 'incense']),
    ('قهوة', ['قهوة', 'القهوة', 'coffee']),
    ('نعناع', ['نعناع', 'النعناع', 'mint']),
    ('تونكا', ['تونكا', 'التونكا', 'tonka']),
    ('حليب', ['حليب', 'الحليب', 'milk']),
    ('لوز', ['لوز', 'اللوز', 'almond']),
    ('جوز الهند', ['جوز الهند', 'coconut']),
    ('شاي', ['شاي أخضر', 'الشاي', 'tea']),
    ('طحلب', ['طحلب', 'طحالب', 'moss', 'oakmoss']),
]


def extract_short_ingredients(ingredients_combined, desc_clean, name):
    """استخلاص قائمة مكونات مختصرة (أهم 8 مكونات)"""
    # بحث في النص الكامل — المكونات تظهر في سطور الافتتاحية/القلب/القاعدة
    search_text = f"{name} {ingredients_combined} {desc_clean[:2000]}".lower()

    found = []
    for display_name, search_forms in KNOWN_INGREDIENTS:
        if display_name in found:
            continue
        for form in search_forms:
            if form.lower() in search_text:
                found.append(display_name)
                break
        if len(found) >= 8:
            break

    return '، '.join(found) if found else ''


# ═══════════════════════════════════════════════════════════
#  المعالجة الرئيسية
# ═══════════════════════════════════════════════════════════

def load_excel_data():
    """قراءة ملف Excel وإرجاع dict بمفتاح الاسم"""
    try:
        import openpyxl
    except ImportError:
        print('❌ مطلوب تثبيت openpyxl: pip install openpyxl')
        sys.exit(1)

    print(f'📂 جاري قراءة Excel: {EXCEL_FILE.name}...')
    wb = openpyxl.load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[2]]
    print(f'   الأعمدة: {len(headers)}')

    # فهرس الأعمدة المهمة
    col_map = {}
    for i, h in enumerate(headers):
        if h == 'أسم المنتج':
            col_map['name'] = i
        elif h == 'تصنيف المنتج':
            col_map['category'] = i
        elif h == 'الوصف':
            col_map['desc'] = i
        elif h == 'الماركة':
            col_map['brand'] = i
        elif h == 'سعر المنتج':
            col_map['price'] = i

    print(f'   أعمدة مهمة: {list(col_map.keys())}')

    excel_products = {}
    row_count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[col_map['name']] if col_map.get('name') is not None else None
        if not name or not isinstance(name, str):
            continue

        name = name.strip()
        category = str(row[col_map.get('category', 0)] or '').strip()
        desc = str(row[col_map.get('desc', 0)] or '').strip()
        brand = str(row[col_map.get('brand', 0)] or '').strip()

        # استخلاص المكونات
        ingredients = extract_ingredients_from_desc(desc)
        desc_clean = strip_html(desc)

        # مكونات مختصرة
        short_ing = extract_short_ingredients(
            ingredients['combined'], desc_clean, name
        )

        # العائلة العطرية
        scent_family = detect_scent_family(
            name, category, ingredients['combined'], desc_clean
        )

        # الجنس
        gender = detect_gender_from_category(category)

        excel_products[name] = {
            'category': category,
            'scent_family': scent_family,
            'ingredients': short_ing,
            'ingredients_full': ingredients,
            'gender_from_cat': gender,
            'brand_excel': brand,
        }

        row_count += 1

    wb.close()
    print(f'   ✅ تم قراءة {row_count} منتج من Excel')
    return excel_products


def normalize_name(name):
    """تطبيع اسم المنتج للمقارنة المرنة"""
    n = name.strip()
    # إزالة المسافات الزائدة
    n = re.sub(r'\s+', ' ', n)
    # إزالة علامات الترقيم للمقارنة
    n = re.sub(r'[()（）\[\]【】]', '', n)
    return n


def enrich_catalog():
    """الدالة الرئيسية: إثراء catalog.json"""

    # 1. تحميل بيانات Excel
    excel_data = load_excel_data()

    # 2. تحميل الكتالوج الحالي
    print(f'\n📦 جاري تحميل الكتالوج: {CATALOG_FILE.name}...')
    with open(CATALOG_FILE, 'r', encoding='utf-8') as f:
        catalog = json.load(f)
    print(f'   المنتجات الحالية: {len(catalog)}')

    # 3. نسخة احتياطية
    print(f'💾 إنشاء نسخة احتياطية: {BACKUP_FILE.name}')
    with open(BACKUP_FILE, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    # 4. بناء فهرس بحث مرن (اسم مطبّع → بيانات Excel)
    excel_index = {}
    for name, data in excel_data.items():
        norm = normalize_name(name)
        excel_index[norm] = data
        # أيضاً بدون "تستر" و"عطر" في البداية
        for prefix in ['تستر ', 'عطر ', 'تيستر ']:
            if norm.startswith(prefix):
                excel_index[norm[len(prefix):]] = data

    # 5. مطابقة وإثراء
    matched = 0
    enriched_with_ingredients = 0
    enriched_with_category = 0
    unmatched_names = []

    for product in catalog:
        pname = product.get('name', '').strip()
        pname_norm = normalize_name(pname)

        # بحث مباشر
        excel_info = excel_index.get(pname_norm)

        # بحث بدون بادئة
        if not excel_info:
            for prefix in ['تستر ', 'عطر ', 'تيستر ']:
                if pname_norm.startswith(prefix):
                    excel_info = excel_index.get(pname_norm[len(prefix):])
                    if excel_info:
                        break

        # بحث جزئي (أول 30 حرف)
        if not excel_info and len(pname_norm) > 20:
            for ename, edata in excel_index.items():
                if pname_norm[:30] in ename or ename[:30] in pname_norm:
                    excel_info = edata
                    break

        if excel_info:
            matched += 1

            # إضافة التصنيف
            if excel_info['category']:
                product['category'] = excel_info['category']
                enriched_with_category += 1

            # إضافة العائلة العطرية
            product['scent_family'] = excel_info['scent_family']

            # إضافة المكونات المختصرة
            if excel_info['ingredients']:
                product['ingredients'] = excel_info['ingredients']
                enriched_with_ingredients += 1

            # تحديث الجنس إذا كان ناقصاً
            if excel_info['gender_from_cat'] and product.get('g') == 'مشترك':
                product['g'] = excel_info['gender_from_cat']

        else:
            unmatched_names.append(pname)
            # حتى لو ما تطابق — نحاول نحدد العائلة من الاسم
            product['scent_family'] = detect_scent_family(
                pname, '', '', ''
            )

    # 6. حفظ الكتالوج المُحدّث
    print(f'\n💾 جاري حفظ الكتالوج المُحدّث...')
    with open(CATALOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    # 7. إحصائيات
    print(f'\n{"="*50}')
    print(f'📊 نتائج الإثراء:')
    print(f'   إجمالي المنتجات في الكتالوج: {len(catalog)}')
    print(f'   تمت المطابقة مع Excel:      {matched} ({matched*100//len(catalog)}%)')
    print(f'   أُثري بالتصنيف:              {enriched_with_category}')
    print(f'   أُثري بالمكونات:             {enriched_with_ingredients}')
    print(f'   لم يتطابق:                   {len(unmatched_names)}')

    # عائلات الروائح
    family_counts = {}
    for p in catalog:
        sf = p.get('scent_family', 'unknown')
        family_counts[sf] = family_counts.get(sf, 0) + 1

    print(f'\n🌸 توزيع العوائل العطرية:')
    for fam in FAMILY_PRIORITY:
        count = family_counts.get(fam, 0)
        bar = '█' * (count // 50)
        print(f'   {fam:12s}: {count:5d} {bar}')

    # عينات
    print(f'\n📝 عينات من المنتجات المُثراة:')
    samples = [p for p in catalog if p.get('ingredients')][:5]
    for s in samples:
        print(f'   🔹 {s["name"][:50]}')
        print(f'      التصنيف: {s.get("category", "—")[:60]}')
        print(f'      العائلة: {s.get("scent_family", "—")}')
        print(f'      المكونات: {s.get("ingredients", "—")[:80]}')
        print()

    # أول 10 غير متطابقين
    if unmatched_names:
        print(f'\n⚠️  أول 10 منتجات لم تتطابق:')
        for n in unmatched_names[:10]:
            print(f'   ❌ {n[:70]}')

    print(f'\n{"="*50}')
    print(f'✅ تم تحديث {CATALOG_FILE.name} بنجاح!')
    print(f'💾 نسخة احتياطية: {BACKUP_FILE.name}')

    return {
        'total': len(catalog),
        'matched': matched,
        'with_ingredients': enriched_with_ingredients,
        'with_category': enriched_with_category,
        'families': family_counts,
    }


if __name__ == '__main__':
    enrich_catalog()
